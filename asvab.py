import requests
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import os
import json

# Get Bearer token from environment variable or use a default
BEARER_TOKEN = os.environ.get('ASVAB_BEARER_TOKEN', 'MGKVnam26x6dapNgwi48YKwX')

# Your Memberful endpoint and headers
url = 'https://asvabadvantage.memberful.com/api/graphql'
headers = {
    'Content-Type': 'application/json',
    'Authorization': f'Bearer {BEARER_TOKEN}',
}

# Template for the GraphQL query with pagination
query_template = """
{
  members(first: 100, after: "%s") {
    edges {
      node {
        email
        trackingParams
        totalSpendCents
        subscriptions {
          activatedAt
          orders {
            createdAt
            status
            totalCents
            type
          }
        }
      }
    }
    pageInfo {
      endCursor
      hasNextPage
    }
  }
}
"""

# Initialize tracking dictionaries
all_subscriptions = defaultdict(set)  # Track all subscriptions by month
all_emails = set()  # Track all unique emails
all_months = set()  # Track all months

# Fetch all members using pagination
has_next_page = True
end_cursor = None

while has_next_page:
    # Update the query with the current end_cursor
    query = {"query": query_template % (end_cursor if end_cursor else "")}
    try:
        response = requests.post(url, headers=headers, json=query)
        response.raise_for_status()
        
        data = response.json()
        
        # Process the current page of data
        for member in data['data']['members']['edges']:
            email = member['node']['email']
            subscriptions = member['node']['subscriptions']
            totalSpend = member['node']['totalSpendCents']
            
            if totalSpend > 0:  # Only process paid members
                all_emails.add(email)
                
                # Process all subscriptions
                for subscription in subscriptions:
                    if subscription and subscription.get("activatedAt"):
                        activated_date = datetime.fromtimestamp(int(subscription['activatedAt']))
                        year_month = activated_date.strftime('%Y-%m')
                        all_subscriptions[year_month].add(email)
                        all_months.add(year_month)
                        
                        # Process all orders for continued subscriptions
                        if subscription.get("orders"):
                            for order in subscription["orders"]:
                                if order.get("createdAt"):
                                    order_date = datetime.fromtimestamp(int(order['createdAt']))
                                    order_month = order_date.strftime('%Y-%m')
                                    if order_month != year_month:
                                        all_subscriptions[order_month].add(email)
                                        all_months.add(order_month)
        
        # Prepare for the next iteration
        page_info = data['data']['members']['pageInfo']
        has_next_page = page_info['hasNextPage']
        end_cursor = page_info['endCursor']
        
    except Exception as e:
        print(f"Error processing data: {e}")
        break

# Create Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Subscribers by Month"

# Sort months and emails
sorted_months = sorted(all_months)
sorted_emails = sorted(all_emails)

# Define styles
header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
header_font = Font(bold=True)
checkmark_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

# Write month headers
for col, month in enumerate(sorted_months, start=2):
    cell = ws.cell(row=1, column=col, value=month)
    cell.fill = header_fill
    cell.font = header_font

# Write email header
ws.cell(row=1, column=1, value="Email").fill = header_fill
ws.cell(row=1, column=1, value="Email").font = header_font

# Add revenue row with subscriber count
ws.cell(row=2, column=1, value="Monthly Revenue ($)").font = header_font
for col, month in enumerate(sorted_months, start=2):
    subscriber_count = len(all_subscriptions[month])
    revenue = subscriber_count * 11
    ws.cell(row=2, column=col, value=revenue)
    ws.cell(row=3, column=col, value=f"({subscriber_count} subscribers)")

# Fill in the data
for row, email in enumerate(sorted_emails, start=4):
    ws.cell(row=row, column=1, value=email)
    for col, month in enumerate(sorted_months, start=2):
        if email in all_subscriptions[month]:
            cell = ws.cell(row=row, column=col, value="✓")
            cell.fill = checkmark_fill

# Adjust column widths
ws.column_dimensions[get_column_letter(1)].width = 35  # Email column
for col in range(2, len(sorted_months) + 2):
    ws.column_dimensions[get_column_letter(col)].width = 15

# Print monthly summary
print("\nMonthly Revenue Summary:")
for month in sorted_months:
    subscriber_count = len(all_subscriptions[month])
    revenue = subscriber_count * 11
    print(f"{month}: ${revenue:.2f} ({subscriber_count} subscribers)")

# Save the workbook
wb.save('subscribers_by_month.xlsx')
print(f"\nExcel file 'subscribers_by_month.xlsx' has been created successfully.")

